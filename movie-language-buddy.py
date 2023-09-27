import translators as ts
import openpyxl


class Language_buddy:
  
  def __init__(self, input_language, output_language, csv_file_path, sheet):
    self.input_language = input_language
    self.output_language = output_language
    self.csv_file_path = csv_file_path
    self.sheet = sheet
    self.csv_dict = {}
  

  def translate(self, input):
    translation = ts.translate_text(input, from_language=self.input_language, to_language=self.output_language)
    print(translation, "\n")

    return translation


  def remove_last(self):
    workbook = openpyxl.load_workbook(self.csv_file_path)
    sheet = workbook[self.sheet]
    last_row = sheet.max_row
    row_data = [sheet.cell(row=last_row, column=col).value for col in range(1, sheet.max_column + 1)]
    print(f"removed {row_data[0]} successfully", "\n")
    sheet.delete_rows(last_row)
    workbook.save(self.csv_file_path)

    return
  

  def show_last(self, number): 
    workbook = openpyxl.load_workbook(self.csv_file_path)
    sheet = workbook[self.sheet]
    total_rows = sheet.max_row
    start_row = max(total_rows - number + 1, 1)

    for row_number in range(start_row, total_rows + 1):
      row_data = [sheet.cell(row=row_number, column=col).value for col in range(1, sheet.max_column + 1)]
      print(f"{row_data[0]} - {row_data[1]}")

    print("")
    return


  def push_to_csv(self, input, translation):
    workbook = openpyxl.load_workbook(self.csv_file_path)
    sheet = workbook[self.sheet]
    row_data = [input, translation]
    sheet.append(row_data)
    workbook.save(self.csv_file_path)
    
    return
  


def extract_int(input_string):
  extracted_digits = ""
  for char in reversed(input_string):
    if char.isdigit():
      extracted_digits = char + extracted_digits
    else:
      break
  if extracted_digits:
    extracted_integer = int(extracted_digits)
    return extracted_integer
  else:
    return None


def main():
  input_language = input("input language: ")
  output_language = input("output language: ")
  csv_file_path = input("csv file: ")
  sheet = input("show: ")
  print("\n")

  language_buddy = Language_buddy(input_language, output_language, csv_file_path, sheet)

  while True:
    prompt = input("")
    if prompt == "save":
      language_buddy.save_to_csv()
    elif prompt == "exit":
      break
    elif "show" in prompt:
      number = extract_int(prompt)
      language_buddy.show_last(number)
    elif "remove" in prompt:
      language_buddy.remove_last()
    else:
      translation = language_buddy.translate(prompt)
      language_buddy.push_to_csv(prompt, translation)
      
    

main()